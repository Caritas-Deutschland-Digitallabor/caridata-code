import pathlib

from openpyxl import load_workbook

SOURCE_PATH = pathlib.Path(__file__).parent.parent.parent.parent

SCHEMA_DEFINITION_PATH = (
    SOURCE_PATH
    / "data"
    / "raw"
    / "Schnittstellendefinition_der_GSDA_stand_2023__1_.xlsx"
)


def get_row_colors(schema_name: str, color_mapper: dict[str, bool]):
    wb = load_workbook(filename=SCHEMA_DEFINITION_PATH, data_only=True)
    ws = wb[schema_name]
    colors: dict[int, bool] = {}
    index = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        color_ = str(cell.fill.fgColor.index)
        color = color_mapper.get(color_, False)
        colors[index] = color
        index += 1
    return colors


def map_colors(schemas: list[str]):
    color_mapper: dict[str, bool] = {
        "13": True,
        "11": False,
        "FFFFFF00": True,
        "42": False,
        "43": True,
    }

    schema_color_mapper = {}

    for schema_name in schemas:
        colors_per_schema = get_row_colors(
            schema_name=schema_name, color_mapper=color_mapper
        )
        schema_color_mapper[schema_name] = colors_per_schema

    return schema_color_mapper
